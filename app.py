import base64
import csv
import hashlib
import io
import json
import os
import re
import shutil
import sqlite3
import tempfile
import traceback
import zipfile
from collections import OrderedDict
from datetime import datetime
from io import BytesIO

import cv2
import fitz
import numpy as np
from PIL import Image
from PyPDF2 import PdfReader, PdfWriter
from flask import Flask, Response, jsonify, render_template, request, send_file, send_from_directory, url_for
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from werkzeug.utils import secure_filename

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
USERDATA_DIR = os.path.join(BASE_DIR, ".userdata")
UPLOADS_FOLDER = os.path.join(USERDATA_DIR, "uploads")
DB_FILE = os.path.join(USERDATA_DIR, "users.db")

os.makedirs(UPLOADS_FOLDER, exist_ok=True)

DEFAULT_USERNAME = "local"

app = Flask(__name__, template_folder=os.path.join(BASE_DIR, "templates"), static_folder=os.path.join(BASE_DIR, "static"))
app.config["UPLOADS_FOLDER"] = UPLOADS_FOLDER
CORS(app)

# Template image cache
template_image_cache = OrderedDict()
MAX_CACHE_SIZE = 10

# Question paper temporary storage
question_paper_storage = {}


def compute_username_hash(username):
    normalized = str(username or "").strip().lower()
    if not normalized:
        return None
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()


def get_db_connection():
    conn = sqlite3.connect(DB_FILE, timeout=15)
    conn.row_factory = sqlite3.Row
    return conn


def initialize_database():
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            username_hash TEXT UNIQUE,
            email TEXT,
            email_hash TEXT,
            password BLOB NOT NULL
        )
        """
    )

    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS scan_batches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            batch_datetime TEXT NOT NULL,
            paper_count INTEGER NOT NULL,
            batch_name TEXT,
            roi_configs TEXT,
            answer_key TEXT,
            question_paper_path TEXT,
            notes TEXT,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
        """
    )

    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS scanned_results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_id INTEGER NOT NULL,
            student_id TEXT,
            score INTEGER NOT NULL,
            total_questions INTEGER NOT NULL,
            answers_json TEXT,
            uncertainties_json TEXT,
            original_image_path TEXT,
            visualization_path TEXT,
            FOREIGN KEY (batch_id) REFERENCES scan_batches (id)
        )
        """
    )

    conn.commit()
    conn.close()


def ensure_default_user():
    username_hash = compute_username_hash(DEFAULT_USERNAME)
    if not username_hash:
        return

    conn = get_db_connection()
    cursor = conn.cursor()
    existing = cursor.execute("SELECT id FROM users WHERE username_hash = ?", (username_hash,)).fetchone()
    if not existing:
        cursor.execute(
            "INSERT INTO users (username, username_hash, email, email_hash, password) VALUES (?, ?, ?, ?, ?)",
            (DEFAULT_USERNAME, username_hash, None, None, b""),
        )
        conn.commit()
    conn.close()


def get_user_id_by_username(cursor, username):
    username_hash = compute_username_hash(username)
    if not username_hash:
        return None
    return cursor.execute("SELECT id FROM users WHERE username_hash = ?", (username_hash,)).fetchone()


def cleanup_template_cache():
    if len(template_image_cache) > MAX_CACHE_SIZE:
        template_image_cache.popitem(last=False)


def normalize_min_pixel_threshold(raw_threshold, roi_configs):
    try:
        threshold_value = float(raw_threshold)
    except (TypeError, ValueError):
        return 300

    if threshold_value <= 0:
        return 300

    if threshold_value <= 1.0:
        return max(1, int(round(threshold_value * 500)))

    return max(1, int(threshold_value))


class BubbleSheetProcessor:
    def __init__(self):
        self.min_pixel_threshold = 300
        self.uncertainty_margin = 50
        self.id_digit_positions = []

    def extract_student_id(self, thresh_img):
        student_id_digits = []

        if not self.id_digit_positions:
            return "?"

        for digit_column in self.id_digit_positions:
            fill_counts = []
            for (x, y, w, h) in digit_column:
                if y + h <= thresh_img.shape[0] and x + w <= thresh_img.shape[1]:
                    digit_area = thresh_img[y:y + h, x:x + w]
                    filled_pixels = cv2.countNonZero(digit_area)
                    fill_counts.append(filled_pixels)
                else:
                    fill_counts.append(0)

            if not fill_counts:
                student_id_digits.append("?")
                continue

            max_fill = max(fill_counts)
            if max_fill < self.min_pixel_threshold:
                selected_digit = "?"
            else:
                second_max = sorted(fill_counts, reverse=True)[1] if len(fill_counts) > 1 else 0
                fill_diff = max_fill - second_max

                if fill_diff <= self.uncertainty_margin:
                    selected_digit = "?"
                else:
                    try:
                        selected_digit = str(fill_counts.index(max_fill))
                    except ValueError:
                        selected_digit = "?"

            student_id_digits.append(selected_digit)

        return "".join(student_id_digits)

    def detect_bubbles_in_roi_realtime_with_threshold(self, image, roi_coords, grid_cols=5, threshold=0.6):
        try:
            x = max(0, int(roi_coords["x"]))
            y = max(0, int(roi_coords["y"]))
            w = min(int(roi_coords["width"]), image.shape[1] - x)
            h = min(int(roi_coords["height"]), image.shape[0] - y)

            if w <= 0 or h <= 0:
                return []

            roi_img = image[y:y + h, x:x + w]
            gray_roi = cv2.cvtColor(roi_img, cv2.COLOR_BGR2GRAY) if len(roi_img.shape) == 3 else roi_img
            _, thresh_roi = cv2.threshold(gray_roi, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)

            bubbles = []

            contours, _ = cv2.findContours(thresh_roi, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

            total_area = w * h
            expected_bubble_area = total_area / (grid_cols * 5)

            min_area = expected_bubble_area * 0.2
            max_area = expected_bubble_area * 3.0

            roi_perimeter = 2 * (w + h)
            min_perimeter = roi_perimeter / (grid_cols * 10)
            max_perimeter = roi_perimeter / (grid_cols * 2)

            valid_bubbles = []

            for contour in contours:
                area = cv2.contourArea(contour)
                perimeter = cv2.arcLength(contour, True)

                size_valid = min_area < area < max_area
                perimeter_valid = min_perimeter < perimeter < max_perimeter

                if size_valid or perimeter_valid:
                    if perimeter > 0:
                        circularity = 4 * np.pi * area / (perimeter * perimeter)

                        if area < expected_bubble_area * 0.5:
                            circularity_threshold = 0.2
                        elif area > expected_bubble_area * 2:
                            circularity_threshold = 0.25
                        else:
                            circularity_threshold = 0.3

                        if circularity > circularity_threshold:
                            valid_bubbles.append(contour)

            if valid_bubbles:
                for contour in valid_bubbles:
                    (center_x, center_y), radius = cv2.minEnclosingCircle(contour)

                    mask = np.zeros(thresh_roi.shape, np.uint8)
                    cv2.drawContours(mask, [contour], -1, 255, -1)
                    filled_pixels = cv2.countNonZero(cv2.bitwise_and(thresh_roi, mask))
                    contour_area = cv2.contourArea(contour)

                    fill_percentage = filled_pixels / contour_area if contour_area > 0 else 0

                    is_filled = fill_percentage > threshold

                    global_x = x + int(center_x)
                    global_y = y + int(center_y)

                    bubbles.append({
                        "x": global_x,
                        "y": global_y,
                        "radius": max(int(radius), 4),
                        "is_filled": is_filled,
                        "fill_percentage": fill_percentage,
                        "bubble_size": cv2.contourArea(contour),
                        "row": 0,
                        "col": 0,
                    })

            else:
                bubbles = self._fallback_grid_detection_with_threshold(thresh_roi, x, y, w, h, grid_cols, threshold)

            return bubbles

        except Exception as exc:
            print(f"Error in adaptive bubble detection: {exc}")
            return []

    def _fallback_grid_detection_with_threshold(self, thresh_roi, x, y, w, h, grid_cols, threshold=0.6):
        bubbles = []

        possible_rows = [
            max(1, h // 35),
            max(1, h // 25),
            max(1, h // 20),
        ]

        best_detection = []
        best_confidence = 0

        for grid_rows in possible_rows:
            current_bubbles = []
            confidence_sum = 0

            cell_width = w / grid_cols
            cell_height = h / grid_rows

            for row in range(grid_rows):
                for col in range(grid_cols):
                    cell_center_x = col * cell_width + cell_width / 2
                    cell_center_y = row * cell_height + cell_height / 2

                    sample_radius = min(cell_width, cell_height) * 0.25

                    sample_x1 = max(0, int(cell_center_x - sample_radius))
                    sample_y1 = max(0, int(cell_center_y - sample_radius))
                    sample_x2 = min(w, int(cell_center_x + sample_radius))
                    sample_y2 = min(h, int(cell_center_y + sample_radius))

                    if sample_x2 > sample_x1 and sample_y2 > sample_y1:
                        sample_area = thresh_roi[sample_y1:sample_y2, sample_x1:sample_x2]
                        filled_pixels = cv2.countNonZero(sample_area)
                        total_pixels = (sample_x2 - sample_x1) * (sample_y2 - sample_y1)

                        fill_percentage = filled_pixels / total_pixels if total_pixels > 0 else 0
                        is_filled = fill_percentage > threshold

                        confidence = 1.0 if (fill_percentage > (threshold + 0.1) or fill_percentage < (threshold - 0.1)) else 0.5
                        confidence_sum += confidence

                        global_x = x + int(cell_center_x)
                        global_y = y + int(cell_center_y)
                        bubble_radius = int(min(cell_width, cell_height) * 0.3)

                        current_bubbles.append({
                            "x": global_x,
                            "y": global_y,
                            "radius": max(bubble_radius, 4),
                            "is_filled": is_filled,
                            "fill_percentage": fill_percentage,
                            "bubble_size": sample_radius * 2,
                            "row": row,
                            "col": col,
                        })

            avg_confidence = confidence_sum / len(current_bubbles) if current_bubbles else 0
            if avg_confidence > best_confidence:
                best_confidence = avg_confidence
                best_detection = current_bubbles

        return best_detection

    def process_omr_with_visual_feedback(self, image, thresh, roi_configs):
        result_image = image.copy()
        bubble_data = []

        for question_idx, choices in enumerate(roi_configs):
            for choice_idx, (x, y, w, h) in enumerate(choices):
                roi = thresh[y:y + h, x:x + w]
                filled_pixels = cv2.countNonZero(roi)

                is_filled = filled_pixels >= self.min_pixel_threshold

                center_x = x + w // 2
                center_y = y + h // 2
                radius = min(w, h) // 3

                color = (0, 255, 0) if is_filled else (0, 0, 255)
                cv2.circle(result_image, (center_x, center_y), radius, color, 2)

                bubble_data.append({
                    "question": question_idx,
                    "choice": choice_idx,
                    "row": question_idx,
                    "col": choice_idx,
                    "is_filled": is_filled,
                    "x": center_x,
                    "y": center_y,
                })

        return result_image, bubble_data

    def process_omr_image(self, image_data, roi_configs, answer_key_data, min_pixel_threshold=300, uncertainty_margin=50):
        try:
            self.min_pixel_threshold = min_pixel_threshold
            self.uncertainty_margin = uncertainty_margin

            if isinstance(image_data, str):
                image_data = base64.b64decode(image_data.split(",")[1] if "," in image_data else image_data)

            pil_image = Image.open(io.BytesIO(image_data))
            image = cv2.cvtColor(np.array(pil_image), cv2.COLOR_RGB2BGR)
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

            _, thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)

            selected_answers = []
            uncertainties = []
            score = 0

            for question_idx, choices in enumerate(roi_configs):
                fill_counts = []

                for (x, y, w, h) in choices:
                    if y + h <= thresh.shape[0] and x + w <= thresh.shape[1]:
                        roi = thresh[y:y + h, x:x + w]
                        filled_pixels = cv2.countNonZero(roi)
                        fill_counts.append(filled_pixels)
                    else:
                        fill_counts.append(0)

                max_fill = max(fill_counts) if fill_counts else 0
                selected_index = fill_counts.index(max_fill) if fill_counts else -1

                if max_fill < self.min_pixel_threshold:
                    selected_answer = "?"
                    reason = "No significant fill detected"
                else:
                    second_max = sorted(fill_counts, reverse=True)[1] if len(fill_counts) > 1 else 0
                    fill_diff = max_fill - second_max

                    if fill_diff <= self.uncertainty_margin:
                        selected_answer = "?"
                        reason = "Multiple bubbles filled"
                    else:
                        selected_answer = chr(65 + selected_index)
                        reason = None

                correct_answer = answer_key_data[question_idx] if question_idx < len(answer_key_data) else "X"

                if selected_answer == "?":
                    uncertainties.append({
                        "question_idx": question_idx,
                        "fill_counts": fill_counts,
                        "reason": reason,
                        "snapshot_url": "",
                    })
                elif correct_answer != "X" and selected_answer == correct_answer:
                    score += 1

                selected_answers.append(selected_answer)

            visual_image, _ = self.process_omr_with_visual_feedback(image, thresh, roi_configs)

            return {
                "score": score,
                "total_questions": len(roi_configs),
                "answers": selected_answers,
                "uncertainties": uncertainties,
                "visualization": visual_image,
            }

        except Exception as exc:
            return {"error": str(exc)}


processor = BubbleSheetProcessor()


def process_single_omr_sheet(image_path, answer_key_list, roi_configs, id_positions, min_px, uncertainty_ratio, batch_id):
    try:
        image = cv2.imread(image_path)
        if image is None:
            return {"error": "Failed to read image"}, []

        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        _, thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)

        processor.min_pixel_threshold = min_px
        processor.id_digit_positions = id_positions

        student_id = processor.extract_student_id(thresh)

        answers = []
        uncertainties = []
        score = 0

        for question_idx, choices in enumerate(roi_configs):
            fill_counts = []

            for (x, y, w, h) in choices:
                if y + h <= thresh.shape[0] and x + w <= thresh.shape[1]:
                    roi = thresh[y:y + h, x:x + w]
                    filled_pixels = cv2.countNonZero(roi)
                    fill_counts.append(filled_pixels)
                else:
                    fill_counts.append(0)

            max_fill = max(fill_counts) if fill_counts else 0

            if max_fill < min_px:
                selected_letter = "?"
                reason = "No significant fill detected"
            else:
                sorted_counts = sorted(fill_counts, reverse=True)
                second_max = sorted_counts[1] if len(sorted_counts) > 1 else 0
                fill_diff = max_fill - second_max

                if fill_diff <= min_px * (1 - uncertainty_ratio):
                    selected_letter = "?"
                    reason = "Multiple bubbles filled"
                else:
                    selected_index = fill_counts.index(max_fill)
                    selected_letter = chr(65 + selected_index)
                    reason = None

            correct_answer = answer_key_list[question_idx] if question_idx < len(answer_key_list) else "X"

            if selected_letter == "?":
                uncertainties.append({
                    "question_idx": question_idx,
                    "fill_counts": fill_counts,
                    "reason": reason,
                    "snapshot_url": "",
                    "image_name": os.path.basename(image_path),
                })
            elif correct_answer != "X" and selected_letter == correct_answer:
                score += 1

            answers.append(selected_letter)

        visualization_dir = os.path.join(UPLOADS_FOLDER, f"batch_{batch_id}_visualizations")
        os.makedirs(visualization_dir, exist_ok=True)
        visualization_filename = f"batch{batch_id}_{os.path.basename(image_path)}"
        visualization_path = os.path.join(visualization_dir, visualization_filename)

        visualization_image, _ = processor.process_omr_with_visual_feedback(image, thresh, roi_configs)
        cv2.imwrite(visualization_path, visualization_image)

        uncertainty_snapshot_dir = os.path.join(UPLOADS_FOLDER, f"batch_{batch_id}_uncertainty")
        os.makedirs(uncertainty_snapshot_dir, exist_ok=True)

        for uncertainty in uncertainties:
            q_idx = uncertainty["question_idx"]
            if q_idx < len(roi_configs):
                choices = roi_configs[q_idx]
                min_x = min(c[0] for c in choices)
                min_y = min(c[1] for c in choices)
                max_x = max(c[0] + c[2] for c in choices)
                max_y = max(c[1] + c[3] for c in choices)

                padding = 10
                min_x = max(0, min_x - padding)
                min_y = max(0, min_y - padding)
                max_x = min(image.shape[1], max_x + padding)
                max_y = min(image.shape[0], max_y + padding)

                snapshot_img = image[min_y:max_y, min_x:max_x]
                snapshot_filename = f"batch{batch_id}_{os.path.basename(image_path)}_q{q_idx + 1}.jpg"
                snapshot_path = os.path.join(uncertainty_snapshot_dir, snapshot_filename)
                cv2.imwrite(snapshot_path, snapshot_img)
                uncertainty["snapshot_url"] = f"/uploads/batch_{batch_id}_uncertainty/{snapshot_filename}"

        return {
            "image_name": os.path.basename(image_path),
            "student_id": student_id,
            "score": score,
            "total_questions": len(roi_configs),
            "answers": answers,
            "uncertainties": uncertainties,
            "visualization_path": os.path.relpath(visualization_path, UPLOADS_FOLDER).replace("\\", "/"),
        }, uncertainties

    except Exception as exc:
        return {"error": str(exc)}, []


@app.route("/")
def omr_page():
    return render_template("OMR.html", username=DEFAULT_USERNAME)


@app.route("/history")
def history_page():
    conn = get_db_connection()
    cursor = conn.cursor()

    user = get_user_id_by_username(cursor, DEFAULT_USERNAME)
    if not user:
        conn.close()
        return render_template("history.html", username=DEFAULT_USERNAME, history=[])

    batches = cursor.execute(
        """
        SELECT id, batch_datetime, paper_count, batch_name, answer_key, question_paper_path
        FROM scan_batches
        WHERE user_id = ?
        ORDER BY batch_datetime DESC
        """,
        (user["id"],),
    ).fetchall()

    history_data = []
    for batch in batches:
        batch_dict = dict(batch)
        batch_id = batch_dict["id"]

        results = cursor.execute(
            """
            SELECT id, student_id, score, total_questions, visualization_path
            FROM scanned_results
            WHERE batch_id = ?
            ORDER BY id
            """,
            (batch_id,),
        ).fetchall()

        web_results = []
        for result in results:
            result_dict = dict(result)
            vis_path = None
            if result_dict.get("visualization_path"):
                vis_path = url_for("uploaded_file", filename=result_dict["visualization_path"])

            web_results.append({
                "id": result_dict["id"],
                "student_id": result_dict["student_id"],
                "score": result_dict["score"],
                "total_questions": result_dict["total_questions"],
                "visualization_url": vis_path,
            })

        history_data.append({
            "batch_info": batch_dict,
            "results": web_results,
        })

    conn.close()
    return render_template("history.html", username=DEFAULT_USERNAME, history=history_data)


@app.route("/uploads/<path:filename>")
def uploaded_file(filename):
    safe_path = os.path.normpath(filename)
    if safe_path.startswith("..") or os.path.isabs(safe_path):
        return "Access denied", 403

    file_path = os.path.join(UPLOADS_FOLDER, safe_path)
    if not os.path.exists(file_path):
        return "File not found", 404

    directory = os.path.dirname(file_path)
    file_name = os.path.basename(file_path)
    return send_from_directory(directory, file_name)


@app.route("/health", methods=["GET"])
def health_check():
    return jsonify({
        "status": "healthy",
        "message": "OMR Lite backend is running",
        "version": "1.0.0",
    })


@app.route("/desktop_capabilities", methods=["GET"])
def desktop_capabilities():
    return jsonify({
        "success": True,
        "desktop_mode": False,
        "native_save_dialog": False,
    })


@app.route("/upload_question_paper", methods=["POST"])
def upload_question_paper():
    try:
        if "question_paper" not in request.files:
            return jsonify({"success": False, "error": "No file provided"}), 400

        file = request.files["question_paper"]
        if file.filename == "":
            return jsonify({"success": False, "error": "Empty filename"}), 400

        filename = secure_filename(file.filename)
        file_id = str(datetime.now().timestamp()).replace(".", "_")
        temp_path = os.path.join(tempfile.gettempdir(), f"qp_{file_id}_{filename}")
        file.save(temp_path)

        question_paper_storage[file_id] = {
            "path": temp_path,
            "filename": filename,
            "uploaded_at": datetime.now(),
        }

        return jsonify({"success": True, "file_id": file_id, "filename": filename}), 200

    except Exception as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


@app.route("/question_paper/<int:batch_id>", methods=["GET"])
def get_question_paper(batch_id):
    conn = get_db_connection()
    cursor = conn.cursor()

    batch = cursor.execute(
        "SELECT question_paper_path FROM scan_batches WHERE id = ?",
        (batch_id,),
    ).fetchone()
    conn.close()

    if not batch or not batch["question_paper_path"]:
        return jsonify({"error": "Question paper not found for this batch"}), 404

    file_path = os.path.join(UPLOADS_FOLDER, batch["question_paper_path"])
    if not os.path.exists(file_path):
        return jsonify({"error": "Question paper file not found"}), 404

    return send_file(file_path, as_attachment=False)


@app.route("/process_template", methods=["POST"])
def process_template():
    try:
        if "template_image" not in request.files:
            return jsonify({"error": "No template image provided"}), 400

        template_file = request.files["template_image"]
        image_data = template_file.read()

        pil_image = Image.open(io.BytesIO(image_data))
        image = cv2.cvtColor(np.array(pil_image), cv2.COLOR_RGB2BGR)

        template_id = "template_" + str(hash(image_data))
        template_image_cache[template_id] = image
        cleanup_template_cache()

        height, width = image.shape[:2]

        return jsonify({
            "success": True,
            "width": width,
            "height": height,
            "template_id": template_id,
            "message": "Template processed successfully",
        })

    except Exception as exc:
        return jsonify({"error": str(exc), "success": False}), 500


@app.route("/detect_bubbles_realtime", methods=["POST"])
def detect_bubbles_realtime():
    try:
        data = request.get_json()

        template_id = data.get("template_id")
        roi_coords = data.get("roi_coords")
        grid_cols = int(data.get("grid_cols", 5))
        threshold = float(data.get("threshold", 0.6))

        if not template_id or template_id not in template_image_cache:
            return jsonify({"error": "Template not found. Please upload template first."}), 400

        if not roi_coords:
            return jsonify({"error": "ROI coordinates not provided"}), 400

        template_image = template_image_cache[template_id]

        bubbles = processor.detect_bubbles_in_roi_realtime_with_threshold(
            template_image, roi_coords, grid_cols, threshold
        )

        if grid_cols and bubbles:
            sorted_bubbles = sorted(bubbles, key=lambda b: (b["y"], b["x"]))

            rows = []
            current_row = [sorted_bubbles[0]]
            y_threshold = sorted_bubbles[0]["radius"] * 2

            for bubble in sorted_bubbles[1:]:
                if abs(bubble["y"] - current_row[0]["y"]) <= y_threshold:
                    current_row.append(bubble)
                else:
                    rows.append(sorted(current_row, key=lambda b: b["x"]))
                    current_row = [bubble]

            if current_row:
                rows.append(sorted(current_row, key=lambda b: b["x"]))

            for row_idx, row in enumerate(rows):
                for col_idx, bubble in enumerate(row):
                    bubble["row"] = row_idx
                    bubble["col"] = col_idx

        return jsonify({
            "success": True,
            "bubbles": bubbles,
            "bubble_count": len(bubbles),
            "threshold": threshold,
        })

    except Exception as exc:
        return jsonify({"error": str(exc), "success": False, "traceback": traceback.format_exc()}), 500


@app.route("/update_bubble_threshold", methods=["POST"])
def update_bubble_threshold():
    try:
        data = request.get_json()

        template_id = data.get("template_id")
        roi_coords = data.get("roi_coords")
        threshold = float(data.get("threshold", 0.6))
        grid_cols = int(data.get("grid_cols", 5))

        if not template_id or template_id not in template_image_cache:
            return jsonify({"error": "Template not found"}), 400

        if not roi_coords:
            return jsonify({"error": "ROI coordinates not provided"}), 400

        template_image = template_image_cache[template_id]

        bubbles = processor.detect_bubbles_in_roi_realtime_with_threshold(
            template_image, roi_coords, grid_cols, threshold
        )

        return jsonify({
            "success": True,
            "bubbles": bubbles,
            "threshold": threshold,
            "bubble_count": len(bubbles),
        })

    except Exception as exc:
        return jsonify({"error": str(exc), "success": False, "traceback": traceback.format_exc()}), 500


@app.route("/validate_roi", methods=["POST"])
def validate_roi():
    try:
        data = request.get_json()
        roi_configs = data.get("roi_configs", [])

        if not roi_configs:
            return jsonify({"error": "No ROI configurations provided"}), 400

        total_questions = len(roi_configs)

        return jsonify({
            "success": True,
            "total_questions": total_questions,
            "message": f"ROI validation successful. {total_questions} questions detected.",
        })

    except Exception as exc:
        return jsonify({"error": str(exc), "success": False}), 500


@app.route("/scan", methods=["POST"])
def scan_bubble_sheet():
    try:
        scan_files = request.files.getlist("scan_images")
        username = request.form.get("username") or DEFAULT_USERNAME
        batch_name = request.form.get("batch_name", "").strip()
        roi_configs_str = request.form.get("roi_configs")
        id_digit_positions_str = request.form.get("id_digit_positions", "[]")
        answer_key_str = request.form.get("answer_key")
        raw_min_pixel_threshold = request.form.get("min_pixel_threshold", 300)
        uncertainty_ratio = float(request.form.get("uncertainty_ratio", 0.80))
        question_paper_id = request.form.get("question_paper_id")

        if not all([scan_files, username, roi_configs_str, answer_key_str]):
            return jsonify({"success": False, "error": "Missing required data"}), 400

        roi_configs = json.loads(roi_configs_str)
        id_digit_positions = json.loads(id_digit_positions_str)
        min_pixel_threshold = normalize_min_pixel_threshold(raw_min_pixel_threshold, roi_configs)

        try:
            answer_key_data = json.loads(answer_key_str)
            if not isinstance(answer_key_data, list):
                return jsonify({"success": False, "error": "Answer key must be a JSON array"}), 400

            question_count = len(roi_configs)
            answer_key_list = ["X"] * question_count

            for item in answer_key_data:
                if isinstance(item, dict) and "question" in item and "answer" in item:
                    try:
                        q_num = int(item["question"]) - 1
                        if 0 <= q_num < question_count:
                            answer_key_list[q_num] = item["answer"].strip().upper()
                    except ValueError:
                        continue

            answer_key = answer_key_list
        except json.JSONDecodeError:
            return jsonify({"success": False, "error": "Invalid JSON format for answer key"}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        user = get_user_id_by_username(cursor, username)
        if not user:
            conn.close()
            return jsonify({"error": f"User '{username}' not found."}), 404

        user_id = user["id"]
        batch_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if not batch_name:
            batch_name = f"Batch_{batch_datetime.replace(':', '-').replace(' ', '_')}"

        sanitized_batch_name = secure_filename(batch_name)

        batch_folder = os.path.join(UPLOADS_FOLDER, sanitized_batch_name)
        os.makedirs(batch_folder, exist_ok=True)

        uncertainty_dir = os.path.join(batch_folder, "uncertainty_snapshots")
        visualization_dir = os.path.join(batch_folder, "visualizations")
        os.makedirs(uncertainty_dir, exist_ok=True)
        os.makedirs(visualization_dir, exist_ok=True)

        question_paper_path = None
        if question_paper_id and question_paper_id in question_paper_storage:
            qp_info = question_paper_storage[question_paper_id]
            qp_filename = f"question_paper_{qp_info['filename']}"
            question_paper_path = os.path.join(batch_folder, qp_filename)
            shutil.move(qp_info["path"], question_paper_path)
            del question_paper_storage[question_paper_id]
            question_paper_path = os.path.join(sanitized_batch_name, qp_filename)

        cursor.execute(
            """
            INSERT INTO scan_batches (user_id, batch_datetime, paper_count, batch_name, roi_configs, answer_key, question_paper_path)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (user_id, batch_datetime, len(scan_files), batch_name, roi_configs_str, answer_key_str, question_paper_path),
        )

        batch_id = cursor.lastrowid

        all_initial_results = []
        uncertain_review_list = []
        processed_count = 0

        for scan_file in scan_files:
            original_filename = secure_filename(scan_file.filename)
            temp_path = os.path.join(tempfile.gettempdir(), original_filename)
            scan_file.save(temp_path)

            initial_result, uncertainties = process_single_omr_sheet(
                temp_path,
                answer_key,
                roi_configs,
                id_digit_positions,
                min_pixel_threshold,
                uncertainty_ratio,
                batch_id,
            )

            if "error" in initial_result:
                student_id = "Unknown"
                final_filename = f"{student_id}_{original_filename}"
            else:
                student_id = initial_result.get("student_id", "Unknown")

                if student_id in ("?", "Unknown"):
                    processed_count += 1
                    student_id = f"Sheet_{processed_count}"

                clean_student_id = re.sub(r"[^\w\-]", "_", student_id)
                file_extension = os.path.splitext(original_filename)[1]
                final_filename = f"{clean_student_id}{file_extension}"

            image_path = os.path.join(batch_folder, final_filename)
            shutil.copy2(temp_path, image_path)
            os.remove(temp_path)

            if "error" not in initial_result and initial_result.get("visualization_path"):
                vis_filename = os.path.basename(initial_result["visualization_path"])
                vis_extension = os.path.splitext(vis_filename)[1]
                new_vis_filename = f"{clean_student_id}_visualization{vis_extension}"
                new_vis_path = os.path.join(visualization_dir, new_vis_filename)

                old_vis_path = os.path.join(UPLOADS_FOLDER, initial_result["visualization_path"])
                if os.path.exists(old_vis_path):
                    shutil.move(old_vis_path, new_vis_path)
                    initial_result["visualization_path"] = os.path.relpath(new_vis_path, UPLOADS_FOLDER).replace("\\", "/")

            if "error" not in initial_result:
                cursor.execute(
                    """
                    INSERT INTO scanned_results
                    (batch_id, student_id, score, total_questions, answers_json, uncertainties_json, original_image_path, visualization_path)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        batch_id,
                        initial_result["student_id"],
                        initial_result["score"],
                        initial_result["total_questions"],
                        json.dumps(initial_result["answers"]),
                        json.dumps(initial_result["uncertainties"]),
                        image_path,
                        initial_result.get("visualization_path", ""),
                    ),
                )

                initial_result["image_name"] = final_filename
                all_initial_results.append(initial_result)

                if uncertainties:
                    for uncertainty in uncertainties:
                        snapshot_url = uncertainty.get("snapshot_url", "")

                        if snapshot_url:
                            snapshot_filename = os.path.basename(snapshot_url)
                            new_snapshot_filename = f"{clean_student_id}_q{uncertainty['question_idx'] + 1}.jpg"
                            new_snapshot_path = os.path.join(uncertainty_dir, new_snapshot_filename)

                            temp_snapshot_dir = os.path.join(UPLOADS_FOLDER, f"batch_{batch_id}_uncertainty")
                            temp_snapshot_path = os.path.join(temp_snapshot_dir, snapshot_filename)

                            if os.path.exists(temp_snapshot_path):
                                shutil.move(temp_snapshot_path, new_snapshot_path)
                                uncertainty["snapshot_url"] = f"/uploads/{sanitized_batch_name}/uncertainty_snapshots/{new_snapshot_filename}"
                            else:
                                img = cv2.imread(image_path)
                                if img is not None:
                                    q_idx = uncertainty["question_idx"]
                                    if q_idx < len(roi_configs):
                                        choices = roi_configs[q_idx]
                                        min_x = min(c[0] for c in choices)
                                        min_y = min(c[1] for c in choices)
                                        max_x = max(c[0] + c[2] for c in choices)
                                        max_y = max(c[1] + c[3] for c in choices)

                                        padding = 10
                                        min_x = max(0, min_x - padding)
                                        min_y = max(0, min_y - padding)
                                        max_x = min(img.shape[1], max_x + padding)
                                        max_y = min(img.shape[0], max_y + padding)

                                        snapshot_img = img[min_y:max_y, min_x:max_x]

                                        cv2.imwrite(new_snapshot_path, snapshot_img)
                                        uncertainty["snapshot_url"] = f"/uploads/{sanitized_batch_name}/uncertainty_snapshots/{new_snapshot_filename}"
                                else:
                                    uncertainty["snapshot_url"] = ""
                        else:
                            uncertainty["snapshot_url"] = ""

                    uncertain_review_list.extend(uncertainties)

            temp_snapshot_dir = os.path.join(UPLOADS_FOLDER, f"batch_{batch_id}_uncertainty")
            if os.path.exists(temp_snapshot_dir):
                try:
                    shutil.rmtree(temp_snapshot_dir)
                except Exception:
                    pass

            temp_vis_dir = os.path.join(UPLOADS_FOLDER, f"batch_{batch_id}_visualizations")
            if os.path.exists(temp_vis_dir):
                try:
                    shutil.rmtree(temp_vis_dir)
                except Exception:
                    pass

        conn.commit()
        conn.close()

        return jsonify({
            "success": True,
            "message": f"Initial scan complete. {len(uncertain_review_list)} items need manual review.",
            "resolved_min_pixel_threshold": min_pixel_threshold,
            "raw_min_pixel_threshold": str(raw_min_pixel_threshold),
            "batch_id": batch_id,
            "batch_name": batch_name,
            "sanitized_batch_name": sanitized_batch_name,
            "results": all_initial_results,
            "uncertain_review_list": uncertain_review_list,
        })

    except Exception as exc:
        return jsonify({"success": False, "error": str(exc), "traceback": traceback.format_exc()}), 500


@app.route("/finalize_scan", methods=["POST"])
def finalize_scan():
    data = request.get_json()
    batch_id = data.get("batch_id")
    corrections = data.get("corrections")
    answer_key_str = data.get("answer_key")

    if not all([batch_id, corrections is not None, answer_key_str]):
        return jsonify({"success": False, "error": "Missing batch_id, corrections, or answer_key"}), 400

    conn = None
    try:
        answer_key_data = json.loads(answer_key_str)
        if not isinstance(answer_key_data, list):
            return jsonify({"success": False, "error": "Answer key must be a JSON array"}), 400

        answer_key = []
        for item in answer_key_data:
            if isinstance(item, dict) and "answer" in item:
                answer_key.append(item["answer"].strip().upper())
            else:
                answer_key.append("X")

        conn = get_db_connection()
        cursor = conn.cursor()

        first_result = cursor.execute(
            "SELECT answers_json FROM scanned_results WHERE batch_id = ? LIMIT 1",
            (batch_id,),
        ).fetchone()

        if not first_result:
            conn.close()
            return jsonify({"success": False, "error": "No results found for this batch"}), 404

        answers_json = json.loads(first_result["answers_json"])
        question_count = len(answers_json)

        while len(answer_key) < question_count:
            answer_key.append("X")

        corrections_by_image = {}
        for corr in corrections:
            image_name = corr["image_name"]
            corrections_by_image.setdefault(image_name, []).append(corr)

        for image_name, sheet_corrections in corrections_by_image.items():
            secure_name = f"batch{batch_id}_{secure_filename(image_name)}"
            original_result = cursor.execute(
                "SELECT id, answers_json FROM scanned_results WHERE batch_id = ? AND original_image_path LIKE ?",
                (batch_id, f"%{secure_name}%"),
            ).fetchone()

            if not original_result:
                continue

            answers = json.loads(original_result["answers_json"])
            for correction in sheet_corrections:
                q_idx = correction["question_idx"]
                if q_idx < len(answers):
                    answers[q_idx] = correction["corrected_answer"]

            new_score = 0
            for i, ans in enumerate(answers):
                if i < len(answer_key) and answer_key[i] != "X" and ans == answer_key[i]:
                    new_score += 1

            cursor.execute(
                "UPDATE scanned_results SET answers_json = ?, score = ? WHERE id = ?",
                (json.dumps(answers), new_score, original_result["id"]),
            )

        conn.commit()

        final_rows = cursor.execute(
            "SELECT * FROM scanned_results WHERE batch_id = ?",
            (batch_id,),
        ).fetchall()

        final_results_list = []
        for row in final_rows:
            image_name = os.path.basename(row["original_image_path"])
            if image_name.startswith(f"batch{batch_id}_"):
                image_name = image_name[len(f"batch{batch_id}_") :]

            final_results_list.append({
                "image_name": image_name,
                "student_id": row["student_id"],
                "score": row["score"],
                "total_questions": row["total_questions"],
                "answers": json.loads(row["answers_json"]),
                "uncertainties": json.loads(row["uncertainties_json"]),
            })

        conn.close()

        return jsonify({
            "success": True,
            "message": "All results finalized.",
            "final_results": final_results_list,
        })

    except Exception as exc:
        if conn:
            conn.close()
        return jsonify({"success": False, "error": str(exc), "traceback": traceback.format_exc()}), 500


@app.route("/download_excel/<int:batch_id>")
def download_excel(batch_id):
    try:
        answer_key_str = request.args.get("answer_key", "")
        if not answer_key_str:
            return jsonify({"error": "Answer key is required"}), 400

        try:
            answer_key_data = json.loads(answer_key_str)
            if not isinstance(answer_key_data, list):
                return jsonify({"error": "Answer key must be a JSON array"}), 400

            answer_key = []
            for item in answer_key_data:
                if isinstance(item, dict) and "answer" in item:
                    answer_key.append(item["answer"].strip().upper())
                else:
                    answer_key.append("X")
        except json.JSONDecodeError:
            return jsonify({"error": "Invalid JSON format for answer key"}), 400

        conn = get_db_connection()
        batch_info = conn.execute(
            "SELECT batch_name FROM scan_batches WHERE id = ?",
            (batch_id,),
        ).fetchone()

        if not batch_info:
            conn.close()
            return jsonify({"error": "Batch not found"}), 404

        results = conn.execute(
            "SELECT * FROM scanned_results WHERE batch_id = ?",
            (batch_id,),
        ).fetchall()
        conn.close()

        if not results:
            return jsonify({"error": "No results found for this batch"}), 404

        wb = Workbook()

        fills = {
            "correct": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "wrong": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "uncertain": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "excluded": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        }

        max_answers = 0
        for result in results:
            answers = json.loads(result["answers_json"])
            max_answers = max(max_answers, len(answers))

        while len(answer_key) < max_answers:
            answer_key.append("X")

        headers = ["Image", "Student ID"]
        for i in range(max_answers):
            headers.append(f"Q{i + 1}")
        headers.extend(["Score", "Total Questions", "Percentage"])

        ws_all = wb.active
        ws_all.title = "All Students"
        ws_all.append(headers)

        for cell in ws_all[1]:
            cell.font = Font(bold=True)

        for result in results:
            answers = json.loads(result["answers_json"])
            image_name = os.path.basename(result["original_image_path"])
            if image_name.startswith(f"batch{batch_id}_"):
                image_name = image_name[len(f"batch{batch_id}_") :]

            percentage = "0%"
            if result["total_questions"] > 0:
                percentage_val = (result["score"] / result["total_questions"]) * 100
                percentage = f"{percentage_val:.1f}%"

            row_data = [image_name, result["student_id"]]
            for i in range(max_answers):
                row_data.append(answers[i] if i < len(answers) else "")
            row_data.extend([result["score"], result["total_questions"], percentage])

            ws_all.append(row_data)
            row_num = ws_all.max_row

            for col_idx in range(len(answers)):
                if col_idx >= len(answer_key):
                    continue
                correct_answer = answer_key[col_idx]
                answer = answers[col_idx]
                cell = ws_all.cell(row=row_num, column=col_idx + 3)

                if correct_answer == "X":
                    cell.fill = fills["excluded"]
                elif answer == correct_answer:
                    cell.fill = fills["correct"]
                elif answer == "?":
                    cell.fill = fills["uncertain"]
                else:
                    cell.fill = fills["wrong"]

        for column in ws_all.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min((max_length + 2) * 1.2, 50)
            ws_all.column_dimensions[column_letter].width = adjusted_width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        batch_name = batch_info["batch_name"] or f"Batch_{batch_id}"
        safe_batch_name = secure_filename(batch_name)
        filename = f"OMR_Results_{safe_batch_name}_Colored.xlsx"

        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as exc:
        return jsonify({"error": str(exc), "traceback": traceback.format_exc()}), 500


@app.route("/download_csv/<int:batch_id>")
def download_csv(batch_id):
    try:
        answer_key_str = request.args.get("answer_key", "")
        if not answer_key_str:
            return jsonify({"error": "Answer key is required"}), 400

        try:
            answer_key_data = json.loads(answer_key_str)
            if not isinstance(answer_key_data, list):
                return jsonify({"error": "Answer key must be a JSON array"}), 400

            answer_key = []
            for item in answer_key_data:
                if isinstance(item, dict) and "answer" in item:
                    answer_key.append(item["answer"].strip().upper())
                else:
                    answer_key.append("X")
        except json.JSONDecodeError:
            return jsonify({"error": "Invalid JSON format for answer key"}), 400

        conn = get_db_connection()
        results = conn.execute(
            "SELECT * FROM scanned_results WHERE batch_id = ?",
            (batch_id,),
        ).fetchall()
        conn.close()

        if not results:
            return jsonify({"error": "No results found for this batch"}), 404

        buffer = io.StringIO()
        writer = csv.writer(buffer)

        headers = ["Image", "Student ID"] + [f"Q{i + 1}" for i in range(len(answer_key))] + [
            "Score",
            "Total Questions",
        ]
        writer.writerow(headers)

        for result in results:
            answers = json.loads(result["answers_json"])
            image_name = os.path.basename(result["original_image_path"])
            if image_name.startswith(f"batch{batch_id}_"):
                image_name = image_name[len(f"batch{batch_id}_") :]

            formatted_answers = []
            for i, ans in enumerate(answers):
                if i >= len(answer_key):
                    formatted_answers.append("")
                    continue

                correct_ans = answer_key[i]

                if correct_ans == "X":
                    formatted_answers.append(f"[EXCLUDED] {ans}")
                elif ans == correct_ans:
                    formatted_answers.append(f"[CORRECT] {ans}")
                elif ans == "?":
                    formatted_answers.append(f"[UNCERTAIN] {ans}")
                else:
                    formatted_answers.append(f"[WRONG] {ans} (Correct: {correct_ans})")

            row_data = [image_name, result["student_id"]] + formatted_answers + [
                result["score"],
                result["total_questions"],
            ]
            writer.writerow(row_data)

        buffer.seek(0)
        csv_data = buffer.getvalue()
        buffer.close()

        filename = f"OMR_Results_Batch_{batch_id}.csv"
        response = Response(
            csv_data,
            mimetype="text/csv",
            headers={"Content-disposition": f"attachment; filename={filename}"},
        )

        return response

    except Exception as exc:
        return jsonify({"error": str(exc), "traceback": traceback.format_exc()}), 500


@app.route("/get_batches", methods=["GET"])
def get_batches():
    try:
        username = request.args.get("username") or DEFAULT_USERNAME
        if not username:
            return jsonify({"success": False, "error": "Username required"}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        user = get_user_id_by_username(cursor, username)
        if not user:
            conn.close()
            return jsonify({"success": False, "error": "User not found"}), 404

        batches = cursor.execute(
            """
            SELECT id, batch_name, batch_datetime, paper_count
            FROM scan_batches
            WHERE user_id = ?
            ORDER BY batch_datetime DESC
            """,
            (user["id"],),
        ).fetchall()

        conn.close()

        batch_list = []
        for batch in batches:
            batch_list.append({
                "id": batch["id"],
                "name": batch["batch_name"] or f"Batch_{batch['id']}",
                "date": batch["batch_datetime"],
                "paper_count": batch["paper_count"],
            })

        return jsonify({"success": True, "batches": batch_list})

    except Exception as exc:
        return jsonify({"success": False, "error": str(exc), "traceback": traceback.format_exc()}), 500


@app.route("/rename_batch/<int:batch_id>", methods=["POST"])
def rename_batch(batch_id):
    try:
        data = request.get_json() or {}
        new_batch_name = data.get("new_batch_name", "").strip()

        if not new_batch_name:
            return jsonify({"success": False, "error": "Batch name cannot be empty"}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        batch = cursor.execute(
            "SELECT batch_name FROM scan_batches WHERE id = ?",
            (batch_id,),
        ).fetchone()

        if not batch:
            conn.close()
            return jsonify({"success": False, "error": "Batch not found"}), 404

        old_batch_name = batch["batch_name"]
        old_sanitized_name = secure_filename(old_batch_name) if old_batch_name else f"Batch_{batch_id}"
        new_sanitized_name = secure_filename(new_batch_name)

        if old_sanitized_name == new_sanitized_name:
            cursor.execute(
                "UPDATE scan_batches SET batch_name = ? WHERE id = ?",
                (new_batch_name, batch_id),
            )
            conn.commit()
            conn.close()
            return jsonify({
                "success": True,
                "message": f"Batch name updated to '{new_batch_name}'",
                "old_batch_name": old_batch_name,
                "new_batch_name": new_batch_name,
            })

        existing_batch = cursor.execute(
            "SELECT id FROM scan_batches WHERE batch_name = ? AND id != ?",
            (new_batch_name, batch_id),
        ).fetchone()

        if existing_batch:
            conn.close()
            return jsonify({"success": False, "error": "A batch with this name already exists"}), 400

        old_folder_path = os.path.join(UPLOADS_FOLDER, old_sanitized_name)
        new_folder_path = os.path.join(UPLOADS_FOLDER, new_sanitized_name)

        if not os.path.exists(old_folder_path):
            conn.close()
            return jsonify({"success": False, "error": f"Batch folder not found: {old_sanitized_name}"}), 404

        if os.path.exists(new_folder_path):
            conn.close()
            return jsonify({"success": False, "error": f"Target folder already exists: {new_sanitized_name}"}), 400

        os.rename(old_folder_path, new_folder_path)

        cursor.execute(
            "UPDATE scan_batches SET batch_name = ? WHERE id = ?",
            (new_batch_name, batch_id),
        )

        results = cursor.execute(
            "SELECT id, original_image_path, visualization_path FROM scanned_results WHERE batch_id = ?",
            (batch_id,),
        ).fetchall()

        for result in results:
            if result["original_image_path"] and old_sanitized_name in result["original_image_path"]:
                new_path = result["original_image_path"].replace(old_sanitized_name, new_sanitized_name)
                cursor.execute(
                    "UPDATE scanned_results SET original_image_path = ? WHERE id = ?",
                    (new_path, result["id"]),
                )

            if result["visualization_path"] and old_sanitized_name in result["visualization_path"]:
                new_vis_path = result["visualization_path"].replace(old_sanitized_name, new_sanitized_name)
                cursor.execute(
                    "UPDATE scanned_results SET visualization_path = ? WHERE id = ?",
                    (new_vis_path, result["id"]),
                )

        conn.commit()
        conn.close()

        return jsonify({
            "success": True,
            "message": f"Batch renamed successfully from '{old_batch_name}' to '{new_batch_name}'",
            "old_batch_name": old_batch_name,
            "new_batch_name": new_batch_name,
        })

    except Exception as exc:
        return jsonify({"success": False, "error": str(exc), "traceback": traceback.format_exc()}), 500


@app.route("/delete_batch/<int:batch_id>", methods=["DELETE", "POST"])
def delete_batch(batch_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        batch = cursor.execute(
            "SELECT batch_name FROM scan_batches WHERE id = ?",
            (batch_id,),
        ).fetchone()

        if not batch:
            conn.close()
            return jsonify({"success": False, "error": "Batch not found"}), 404

        batch_name = batch["batch_name"] or f"Batch_{batch_id}"
        sanitized_name = secure_filename(batch_name)
        folder_path = os.path.join(UPLOADS_FOLDER, sanitized_name)
        if os.path.exists(folder_path):
            shutil.rmtree(folder_path)

        cursor.execute("DELETE FROM scanned_results WHERE batch_id = ?", (batch_id,))
        cursor.execute("DELETE FROM scan_batches WHERE id = ?", (batch_id,))

        conn.commit()
        conn.close()

        return jsonify({"success": True, "message": f"Batch '{batch_name}' deleted successfully"})

    except Exception as exc:
        return jsonify({"success": False, "error": str(exc), "traceback": traceback.format_exc()}), 500


@app.route("/analyze_questions/<int:batch_id>", methods=["GET"])
def analyze_questions(batch_id):
    try:
        answer_key_str = request.args.get("answer_key", "")
        if not answer_key_str:
            return jsonify({"error": "Answer key required"}), 400

        try:
            answer_key_data = json.loads(answer_key_str)
        except json.JSONDecodeError as exc:
            return jsonify({"error": f"Invalid JSON format: {str(exc)}"}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        results = cursor.execute(
            "SELECT answers_json FROM scanned_results WHERE batch_id = ?",
            (batch_id,),
        ).fetchall()
        conn.close()

        if not results:
            return jsonify({"error": "No results found for this batch"}), 404

        question_stats = {}
        for idx, item in enumerate(answer_key_data):
            try:
                question_number = int(item.get("question", idx + 1))
            except (TypeError, ValueError):
                question_number = idx + 1
            correct_answer = str(item.get("answer", "")).upper()

            if correct_answer == "X":
                continue

            question_stats[question_number - 1] = {
                "question_number": question_number,
                "correct_answer": correct_answer,
                "correct_count": 0,
                "wrong_count": 0,
                "uncertain_count": 0,
                "answer_distribution": {},
            }

        total_students = len(results)

        for result in results:
            try:
                answers = json.loads(result["answers_json"])
                for q_idx, answer in enumerate(answers):
                    if q_idx not in question_stats:
                        continue

                    correct_answer = question_stats[q_idx]["correct_answer"]
                    student_answer = str(answer).upper() if answer else ""

                    if student_answer == "?":
                        question_stats[q_idx]["uncertain_count"] += 1
                    elif student_answer == correct_answer:
                        question_stats[q_idx]["correct_count"] += 1
                    else:
                        question_stats[q_idx]["wrong_count"] += 1

                    if student_answer not in question_stats[q_idx]["answer_distribution"]:
                        question_stats[q_idx]["answer_distribution"][student_answer] = 0
                    question_stats[q_idx]["answer_distribution"][student_answer] += 1
            except Exception:
                continue

        analysis = []
        for q_idx in sorted(question_stats.keys()):
            stats = question_stats[q_idx]
            total_answered = stats["correct_count"] + stats["wrong_count"]

            if total_answered > 0:
                correct_percentage = (stats["correct_count"] / total_answered) * 100
            else:
                correct_percentage = 0

            if correct_percentage >= 80:
                difficulty = "Easy"
            elif correct_percentage >= 50:
                difficulty = "Medium"
            else:
                difficulty = "Hard"

            analysis.append({
                "question_number": stats["question_number"],
                "correct_answer": stats["correct_answer"],
                "correct_count": stats["correct_count"],
                "wrong_count": stats["wrong_count"],
                "uncertain_count": stats["uncertain_count"],
                "correct_percentage": round(correct_percentage, 2),
                "difficulty": difficulty,
                "answer_distribution": stats["answer_distribution"],
            })

        return jsonify({
            "success": True,
            "total_students": total_students,
            "analysis": sorted(analysis, key=lambda x: x["question_number"]),
        })

    except Exception as exc:
        return jsonify({"error": str(exc), "traceback": traceback.format_exc()}), 500


@app.route("/generate_scored_pdf/<int:batch_id>", methods=["GET"])
def generate_scored_pdf(batch_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        batch = cursor.execute(
            "SELECT batch_name, batch_datetime, paper_count FROM scan_batches WHERE id = ?",
            (batch_id,),
        ).fetchone()

        if not batch:
            conn.close()
            return jsonify({"error": "Batch not found"}), 404

        results = cursor.execute(
            "SELECT student_id, score, total_questions, visualization_path FROM scanned_results WHERE batch_id = ? ORDER BY student_id",
            (batch_id,),
        ).fetchall()
        conn.close()

        if not results:
            return jsonify({"error": "No results found for this batch"}), 404

        temp_dir = tempfile.mkdtemp()
        output_pdf_path = os.path.join(temp_dir, f"{secure_filename(batch['batch_name'])}_scored.pdf")

        image_list = []
        for result in results:
            if result["visualization_path"]:
                vis_full_path = os.path.join(UPLOADS_FOLDER, result["visualization_path"])
                if os.path.exists(vis_full_path):
                    image_list.append({
                        "path": vis_full_path,
                        "student_id": result["student_id"],
                        "score": result["score"],
                        "total": result["total_questions"],
                    })

        if not image_list:
            return jsonify({"error": "No visualization images found"}), 404

        pdf_writer = PdfWriter()

        for img_info in image_list:
            img = Image.open(img_info["path"])
            if img.mode != "RGB":
                img = img.convert("RGB")

            temp_img_pdf = os.path.join(temp_dir, f"temp_{img_info['student_id']}.pdf")
            img.save(temp_img_pdf, "PDF", resolution=100.0)

            pdf_reader = PdfReader(temp_img_pdf)
            for page in pdf_reader.pages:
                pdf_writer.add_page(page)

        with open(output_pdf_path, "wb") as output_file:
            pdf_writer.write(output_file)

        return send_file(
            output_pdf_path,
            as_attachment=True,
            download_name=f"{secure_filename(batch['batch_name'])}_scored.pdf",
            mimetype="application/pdf",
        )

    except Exception as exc:
        return jsonify({"error": str(exc), "traceback": traceback.format_exc()}), 500


@app.route("/extract_student_data/<int:batch_id>", methods=["GET"])
def extract_student_data(batch_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        batch = cursor.execute(
            "SELECT batch_name, batch_datetime FROM scan_batches WHERE id = ?",
            (batch_id,),
        ).fetchone()

        if not batch:
            conn.close()
            return jsonify({"error": "Batch not found"}), 404

        results = cursor.execute(
            """
            SELECT student_id, score, total_questions, answers_json, uncertainties_json,
                   original_image_path, visualization_path
            FROM scanned_results
            WHERE batch_id = ?
            ORDER BY student_id
            """,
            (batch_id,),
        ).fetchall()
        conn.close()

        student_data = []
        for result in results:
            percentage = 0
            if result["total_questions"] > 0:
                percentage = (result["score"] / result["total_questions"]) * 100

            answers = json.loads(result["answers_json"]) if result["answers_json"] else []
            uncertainties = json.loads(result["uncertainties_json"]) if result["uncertainties_json"] else []

            uncertain_count = len([a for a in answers if a == "?"])

            student_data.append({
                "student_id": result["student_id"],
                "score": result["score"],
                "total_questions": result["total_questions"],
                "percentage": round(percentage, 2),
                "grade": get_letter_grade(percentage),
                "answers": answers,
                "uncertain_count": uncertain_count,
                "has_uncertainties": uncertain_count > 0,
                "image_name": os.path.basename(result["original_image_path"]),
                "visualization_available": bool(result["visualization_path"]),
            })

        return jsonify({
            "success": True,
            "batch_name": batch["batch_name"],
            "batch_date": batch["batch_datetime"],
            "total_students": len(student_data),
            "student_data": student_data,
            "summary": {
                "total_students": len(student_data),
                "average_score": round(sum(s["percentage"] for s in student_data) / len(student_data), 2) if student_data else 0,
                "students_with_uncertainties": sum(1 for s in student_data if s["has_uncertainties"]),
            },
        })

    except Exception as exc:
        return jsonify({"error": str(exc), "traceback": traceback.format_exc()}), 500


def get_letter_grade(percentage):
    if percentage >= 90:
        return "A"
    if percentage >= 80:
        return "B"
    if percentage >= 70:
        return "C"
    if percentage >= 60:
        return "D"
    return "F"


# ==================== PDF PROCESSING ====================

def align_images(im, im_reference):
    try:
        im1_gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
        im2_gray = cv2.cvtColor(im_reference, cv2.COLOR_BGR2GRAY)

        orb = cv2.ORB_create(5000)

        keypoints1, descriptors1 = orb.detectAndCompute(im1_gray, None)
        keypoints2, descriptors2 = orb.detectAndCompute(im2_gray, None)

        if descriptors1 is None or descriptors2 is None:
            return None

        matcher = cv2.DescriptorMatcher_create(cv2.DESCRIPTOR_MATCHER_BRUTEFORCE_HAMMING)
        matches = matcher.match(descriptors1, descriptors2, None)

        matches = sorted(matches, key=lambda x: x.distance)

        num_good_matches = int(len(matches) * 0.15)
        matches = matches[:num_good_matches]

        if len(matches) < 4:
            return None

        points1 = np.zeros((len(matches), 2), dtype=np.float32)
        points2 = np.zeros((len(matches), 2), dtype=np.float32)

        for i, match in enumerate(matches):
            points1[i, :] = keypoints1[match.queryIdx].pt
            points2[i, :] = keypoints2[match.trainIdx].pt

        h, _ = cv2.findHomography(points1, points2, cv2.RANSAC)

        if h is None:
            return None

        height, width = im_reference.shape[:2]
        im_aligned = cv2.warpPerspective(im, h, (width, height), flags=cv2.INTER_LANCZOS4, borderMode=cv2.BORDER_CONSTANT, borderValue=(255, 255, 255))

        return im_aligned

    except Exception:
        return None


@app.route("/rotate_align_pdf", methods=["POST"])
def rotate_align_pdf():
    try:
        if "pdf_file" not in request.files:
            return jsonify({"error": "No PDF file provided"}), 400

        pdf_file = request.files["pdf_file"]
        rotation_angle = int(request.form.get("rotation_angle", 0))
        auto_align = request.form.get("auto_deskew", "true").lower() == "true"

        pdf_bytes = pdf_file.read()
        pdf_document = fitz.open(stream=pdf_bytes, filetype="pdf")

        output_images = []
        temp_dir = tempfile.mkdtemp()

        total_pages = len(pdf_document)

        all_pages = []
        for page_num in range(total_pages):
            page = pdf_document[page_num]
            if rotation_angle != 0:
                page.set_rotation(rotation_angle)

            mat = fitz.Matrix(5, 5)
            pix = page.get_pixmap(matrix=mat)

            img_data = pix.samples
            img = np.frombuffer(img_data, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)

            if pix.n == 4:
                img = cv2.cvtColor(img, cv2.COLOR_RGBA2BGR)
            elif pix.n == 1:
                img = cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)

            all_pages.append(img)

        pdf_document.close()

        if len(all_pages) > 0:
            reference_image = all_pages[0]

            for page_num, img in enumerate(all_pages):
                aligned = False

                if page_num > 0 and auto_align:
                    aligned_img = align_images(img, reference_image)
                    if aligned_img is not None:
                        img = aligned_img
                        aligned = True

                img_filename = f"page_{page_num + 1}.png"
                img_path = os.path.join(temp_dir, img_filename)
                cv2.imwrite(img_path, img)

                output_images.append({
                    "page_number": page_num + 1,
                    "filename": img_filename,
                    "path": img_path,
                    "aligned": aligned,
                })

        zip_path = os.path.join(temp_dir, "aligned_pages.zip")
        with zipfile.ZipFile(zip_path, "w") as zipf:
            for img_info in output_images:
                zipf.write(img_info["path"], img_info["filename"])

        return send_file(
            zip_path,
            as_attachment=True,
            download_name="aligned_pages.zip",
            mimetype="application/zip",
        )

    except Exception as exc:
        return jsonify({"error": str(exc), "traceback": traceback.format_exc()}), 500


@app.route("/convert_to_pdf", methods=["POST"])
def convert_to_pdf():
    try:
        image_files = request.files.getlist("images")
        if not image_files:
            return jsonify({"error": "No images provided"}), 400

        pdf_bytes = BytesIO()
        pdf_document = fitz.open()

        for img_file in image_files:
            img_bytes = img_file.read()
            img = Image.open(BytesIO(img_bytes))

            if img.mode != "RGB":
                img = img.convert("RGB")

            img_pdf_bytes = BytesIO()
            img.save(img_pdf_bytes, format="PDF")
            img_pdf_bytes.seek(0)

            img_pdf = fitz.open(stream=img_pdf_bytes, filetype="pdf")
            pdf_document.insert_pdf(img_pdf)

        pdf_document.save(pdf_bytes)
        pdf_document.close()
        pdf_bytes.seek(0)

        return send_file(
            pdf_bytes,
            as_attachment=True,
            download_name="processed_results.pdf",
            mimetype="application/pdf",
        )

    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.errorhandler(413)
def too_large(error):
    return jsonify({"error": "File too large"}), 413


@app.errorhandler(400)
def bad_request(error):
    return jsonify({"error": "Bad request"}), 400


@app.errorhandler(500)
def internal_error(error):
    return jsonify({"error": "Internal server error"}), 500


if __name__ == "__main__":
    os.makedirs(USERDATA_DIR, exist_ok=True)
    initialize_database()
    ensure_default_user()
    app.run(host="127.0.0.1", port=5001, debug=True)
